#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 Excel 配置表导出为 JSON 文件。

运行方式（必须从 Client/ 目录执行）：
    python Tools/export_config.py

输出目录：Assets/Config/Json/
    GameplaySettings.json
        扁平键值表，供客户端和服务端共用
    ClanSettings.json
        扁平键值表，部分字段会被解析为数组，供客户端和服务端共用
    BuildingAvailability.json
        按大本营等级分组的嵌套结构，供客户端和服务端共用

主要规则：
    - 单元格中若包含“|”，优先按数组解析，例如 "5|10|15" -> [5, 10, 15]
    - 能识别为整数的值导出为 int，能识别为浮点数的值导出为 float，其余保留为字符串
    - Unity 侧可通过编辑器菜单中的 DataConfigMigration 将 JSON 写入 ScriptableObject
    - 服务端可直接把这些 JSON 作为配置源读取

说明：
    这个脚本只负责“Excel -> JSON”这一段数据转换，不负责直接修改 Unity 资源。
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("[Error] Missing openpyxl. Run: pip install openpyxl")
    sys.exit(1)

# -- 路径配置 ------------------------------------------------------------------
XLSX_PATH  = "Assets/Config/GameConfig.xlsx"
OUTPUT_DIR = "Assets/Config/Json"


# -- 类型推断 ------------------------------------------------------------------
def parse_value(raw):
    """把单元格原始值转换为更合适的 Python 类型。

    转换顺序如下：
        1. 空值或空字符串 -> None
        2. 包含“|”       -> 优先解析为整数数组；若失败则退回为字符串数组
        3. 可解析为整数   -> int
        4. 可解析为浮点数 -> float
        5. 其他情况       -> str

    这样做的目的，是让导出的 JSON 尽量保留配置数据的真实语义，
    避免所有内容都以字符串形式落盘，增加后续客户端和服务端的解析负担。
    """
    if raw is None:
        return None
    v = str(raw).strip()
    if not v:
        return None

    # 带“|”的值视为数组配置，例如 "1|2|3"。
    # 这里优先尝试按整数数组解析；如果存在非整数项，则保留为字符串数组。
    if "|" in v:
        parts = [p.strip() for p in v.split("|") if p.strip()]
        try:
            return [int(p) for p in parts]
        except ValueError:
            return parts  # 退回到字符串数组，避免因为单个元素类型不匹配而丢失整行数据

    # 先尝试 int，避免像 "10" 这样的值被误存为 10.0。
    try:
        return int(v)
    except ValueError:
        pass

    # 再尝试 float，用于支持类似 0.5、3.14 这类配置。
    try:
        return float(v)
    except ValueError:
        pass

    return v


# -- 工作表导出函数 -------------------------------------------------------------
def export_key_value_sheet(ws):
    """把 Key / Value / Description 结构的工作表导出为字典。

    约定：
        - 第 1 行是表头，不参与导出
        - 第 1 列是配置键名 Key
        - 第 2 列是配置值 Value
        - 第 3 列通常是说明 Description，本脚本不参与处理

    导出结果中：
        - JSON 的键来自 Key 列
        - JSON 的值来自 Value 列，并通过 parse_value 自动推断类型

    该函数适用于结构简单、字段不需要嵌套的配置表。
    """
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0] if len(row) > 0 else None
        val = row[1] if len(row) > 1 else None
        if not key:
            continue
        key = str(key).strip()
        parsed = parse_value(val)
        if parsed is not None:
            result[key] = parsed
    return result


def export_building_availability(ws):
    """导出建筑可用性配置，并按大本营等级分组。

    期望表结构：
        - 第 1 列：TownHallLevel
        - 第 2 列：BuildingID
        - 第 3 列：Count
        - 第 4 列：MaxLevel

    输出格式：
        {
            "townHallConfigs": [
                {
                    "townHallLevel": 1,
                    "buildings": [ ... ]
                },
                ...
            ]
        }

    顶层键名必须使用 "townHallConfigs"，这是为了与 Unity 侧
    BuildingAvailabilitySO.townHallConfigs 字段保持一致，便于
    JsonUtility.FromJsonOverwrite 直接写入 ScriptableObject。
    """
    # 先按大本营等级聚合到字典中，最后再统一排序输出，
    # 这样可以兼容 Excel 中同一 TH 的多行建筑记录。
    th_map = {}  # { th_level(int): { "townHallLevel": int, "buildings": [...] } }

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        try:
            th_level = int(row[0])
            building_id = str(row[1]).strip()
            count = int(row[2])
            max_level = int(row[3])
        except (TypeError, ValueError, IndexError):
            continue  # 跳过格式错误的行，避免单行脏数据中断整个导出流程

        if th_level not in th_map:
            th_map[th_level] = {"townHallLevel": th_level, "buildings": []}

        th_map[th_level]["buildings"].append({
            "id": building_id,
            "count": count,
            "maxLevel": max_level
        })

    # 按 TH 等级升序导出，保证输出稳定，便于测试和版本对比。
    configs = [th_map[k] for k in sorted(th_map.keys())]
    return {"townHallConfigs": configs}


# -- 主流程 --------------------------------------------------------------------
def main():
    if not os.path.exists(XLSX_PATH):
        print(f"[Error] Excel file not found: {XLSX_PATH}")
        print("        请确认当前命令是在 Client/ 目录下执行的。")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # data_only=True 表示如果单元格中有公式，则读取公式计算后的结果，
    # 而不是像 "=A1+B1" 这样的公式文本。
    # read_only=True 则表示只读打开，适合导出场景，内存占用也更低。
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    exports = [
        ("GameplaySettings",     export_key_value_sheet),
        ("ClanSettings",         export_key_value_sheet),
        ("BuildingAvailability", export_building_availability),
    ]

    for sheet_name, export_fn in exports:
        if sheet_name not in wb.sheetnames:
            print(f"[Warning] Sheet not found: {sheet_name}")
            continue

        # 每张表都交给对应的导出函数处理，这里统一负责调度和写文件。
        data     = export_fn(wb[sheet_name])
        out_path = os.path.join(OUTPUT_DIR, f"{sheet_name}.json")

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        if sheet_name == "BuildingAvailability":
            count = len(data.get("townHallConfigs", []))
            print(f"[Export] {out_path}  ({count} town hall levels)")
        else:
            print(f"[Export] {out_path}  ({len(data)} entries)")

    wb.close()
    print(f"\n[Done] JSON files written to {OUTPUT_DIR}/")
    print("       Unity: 菜单 Tools -> Config -> Generate SO Assets from Data.cs")


if __name__ == "__main__":
    main()
